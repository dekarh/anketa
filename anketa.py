# -*- coding: utf-8 -*-
# Берем из монго, пропускаем через матрицу (excel), сортируем, печатаем результат

import sys, argparse
from _datetime import datetime, timedelta, date
import os
from mysql.connector import MySQLConnection, Error
from collections import OrderedDict
import openpyxl
from pymongo import MongoClient
from lib import read_config, fine_phone
import psycopg2

QUESTIONS = ['financial_state','financial_strategy','savings_strategy','savings_state','savings_target',
             'savings_method','savings_insurance','personal_credit','personal_credit_debt','personal_accounting',
             'savings_safest_method','savings_profitable_method','product_analytics','mlm_awareness','insurance_state',
             'pension_awareness','pension_contract','pension_payments_awareness','information_reliable_source',
             'secured_rights','secured_rights_police','financial_education_level','financial_education_sufficient',
             'financial_education_update','education_conference','education_conference_theme',
             'information_source_list','financial_subject_school']

cfg = read_config(filename='anketa.ini', section='postgresql')
conn = psycopg2.connect(**cfg)
cursor = conn.cursor()
cursor.execute('select ac.id, ac.lastname, ac.name, ac.middlename, di.title from account as ac '
               'left join division as di on ac.division_id = di.id;')
agents = {}
for row in cursor:
    agents[row[0]] = row
agents_descriptions = {}
for i, description in enumerate(cursor.description):
    agents_descriptions[description.name] = i
cursor.close()
conn.close()

wb = openpyxl.load_workbook(filename='agents.xlsx', read_only=True)
ws = wb[wb.sheetnames[0]]
agents_xls_city = {}
agents_xls_firm = {}
for i, row in enumerate(ws):
    if i > 1 and row[2].value:
        agents_xls_city[row[2].value.upper()] = row[0].value
        agents_xls_firm[row[2].value.upper()] = row[1].value

wb = openpyxl.load_workbook(filename='key.xlsx', read_only=True)
ws = wb[wb.sheetnames[0]]
name_of_categories = []
for i, row in enumerate(ws):
    if i > 0:
        break
    for j, cell in enumerate(row):
        if j < 4:
            continue
        name_of_categories.append(cell.value)

categories = {}
for i, row in enumerate(ws):
    if i < 1:
        continue
    for j, cell in enumerate(row):
        if j == 0:
            question = QUESTIONS[int(cell.value) - 1]
        elif j == 2:
            n_answer = 100 * int(cell.value)
        elif j == 3:
            answer = cell.value
        elif j > 3:
            if cell.value:
                if categories.get(question, None):
                    if categories[question].get(n_answer, None):
                        categories[question][n_answer][name_of_categories[j - 4]] = name_of_categories[j - 4]
                    else:
                        categories[question][n_answer] = {name_of_categories[j - 4]: name_of_categories[j - 4]}
                else:
                    categories[question] = {n_answer: {name_of_categories[j - 4]: name_of_categories[j - 4]}}
        else:
            continue


# подключаемся к серверу
cfg = read_config(filename='anketa.ini', section='Mongo')
conn = MongoClient('mongodb://' + cfg['user'] + ':' + cfg['password'] + '@' + cfg['ip'] + ':' + cfg['port'] + '/'
                   + cfg['db'])
# выбираем базу данных
db = conn.saturn_v

# выбираем коллекцию документов
colls = db.Provider_Finfort_Poll

anketes = []
for coll in colls.find():
    fio = coll['passport_lastname'] + ' ' + coll['passport_name'] + ' ' + coll['passport_middlename']
    anketes.append({'fio': fio, 'phone': fine_phone(coll['personal_phone']), 'created': coll['created_date'],
                    'city': coll['city'], 'owner_id': coll['owner_id'], 'question_list': coll['question_list']})
wb_rez = openpyxl.Workbook(write_only=True)
ws_rez = wb_rez.create_sheet('Количество по категориям')
ws_rez.append(['Город Агента', 'Юр.лицо Агента', 'ФИО Агента', 'Подразделение', 'Город Клиента', 'ФИО Клиента',
               'Телефон Клиента', 'Дата и время создания', 'Категории ->'])
for ankete in reversed(anketes):
    sum_categories = {}
    for question in ankete['question_list']:
        n_answer = ankete['question_list'][question]
        if categories.get(question, None):
            if str(type(n_answer)).find('list') > -1:
                for n_answer_i in n_answer:
                    if categories[question].get(n_answer_i, None):
                        for category in categories[question][n_answer_i]:
                            if sum_categories.get(category, None):
                                sum_categories[category] += 1
                            else:
                                sum_categories[category] = 1
            else:
                if categories[question].get(n_answer, None):
                    for category in categories[question][n_answer]:
                        if sum_categories.get(category, None):
                            sum_categories[category] += 1
                        else:
                            sum_categories[category] = 1
    sum_categories_sorted = OrderedDict(sorted(sum_categories.items(), key=lambda t: t[1],reverse=True))
    sum_categories_sorted4print = ''
    fio = (agents[ankete['owner_id']][agents_descriptions['lastname']] + ' ' +
           agents[ankete['owner_id']][agents_descriptions['name']] + ' ' +
           agents[ankete['owner_id']][agents_descriptions['middlename']]).upper()
    rez_string = [agents_xls_city.get(fio,''), agents_xls_firm.get(fio,''), fio,
                  agents[ankete['owner_id']][agents_descriptions['title']],
                  ankete['city'], ankete['fio'], ankete['phone'], ankete['created']]
    for sum_category in sum_categories_sorted:
        rez_string.append(sum_category + ': ' + str(sum_categories_sorted[sum_category]))
    ws_rez.append(rez_string)
wb_rez.save('rez.xlsx')